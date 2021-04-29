# -*- coding: utf-8 -*-
"""
Created on Tue Apr 20 21:59:02 2021

@author: maria
"""
import openpyxl
import numpy as np

class Bin(object):
    """ This class implements the bin object """
    def __init__(self):
        self.items = []  #the list that is going to contain the items that fit in the bin
        self.total_weight = 0  

    def add_item(self, item): #this function adds an item in the bin
        self.items.append(item)
        self.total_weight += item


def read(f):
    """This function reads a given file and returns a tuple 
    that has the list containing the items and an integer=the bin capacity"""
    book = openpyxl.load_workbook(f)  
    sheet = book.active      
    items_list =[]
    n = sheet.cell(row = 2, column = 1).value
    c = sheet.cell(row = 2, column = 2).value
    for i in range(2,n+2):  
        cell_obj = sheet.cell(row = i, column = 3)
        items_list.append(cell_obj.value) 
    return items_list,c  


def ffa(items_list, bin_capacity):
    """Implements the First Fit Algorithm"""
    bins =[]
    randomised_np_list = np.random.permutation(items_list) # list containing initial items in a random order
    items_list = randomised_np_list.tolist() 
    
    for item in items_list:
        # foeach item we search if there's an open bin where it can fit
        for bin in bins:
            if bin.total_weight + item <= bin_capacity: #if it fits
                bin.add_item(item)  #we add the item in the bin
                break
        else:
            # there is no open bin where the item can fit
            #so we open a new bin and add the item in it
            bin = Bin()
            bin.add_item(item)
            bins.append(bin)

    return bins

def ffda(items_list, bin_capacity):
    """Implements the First Fit Decreasing Algorithm"""
    decreased_list = sorted(items_list,reverse=True)  #sorts the items list in a decreasing order
    bins =[]
    for item in decreased_list:
        # foeach item we search if there's an open bin where it can fit
        for bin in bins:
            if bin.total_weight + item <= bin_capacity: #if it fits
                bin.add_item(item)  #we add the item in the bin
                break
        else:
            # there is no open bin where the item can fit
            #so we open a new bin and add the item in it
            bin = Bin()
            bin.add_item(item)
            bins.append(bin)

    return bins
    


if __name__ == '__main__': 
          
    result=read("N1C1W1_A.xlsx")
    items_list,c = result[0],result[1] 
    bins_ffa = ffa(items_list,c)
    bins_ffda = ffda(items_list,c)
     
    print("ffa solution is using: ",len(bins_ffa), "bins")
    print("ffda solution is using: ",len(bins_ffda), "bins")
    
    # N1C1W1_A_solution = 25
    # N1C1W1_B_solution = 31
    # HARD0_solution = 56
    # while(len(bins_ffa)!=N1C1W1_A_solution):
    #     bins_ffa = ffa(items_list,c)
    #     print("looping..")
    # print("solution found")


   